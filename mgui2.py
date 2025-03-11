import openpyxl
import sqlite3
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

def excel_to_sqlite(excel_configs, db_file='test.db', log_callback=None):
    """
    엑셀 파일의 데이터를 SQLite 데이터베이스로 변환합니다.
    
    Args:
        excel_configs (dict): 엑셀 파일 구성 정보를 담은 딕셔너리
        db_file (str): SQLite 데이터베이스 파일 경로
        log_callback (function): 로그 메시지를 표시하기 위한 콜백 함수
    """
    def log(message):
        if log_callback:
            log_callback(message)
        else:
            print(message)
    
    def clean_column_name(name):
        """컬럼 이름에서 공백과 특수문자를 제거하고 SQLite에 적합한 이름으로 변환합니다."""
        if name is None:
            return ""
        
        # 문자열로 변환
        name = str(name).strip()
        
        # 빈 이름 처리
        if not name:
            return ""
        
        # 공백과 특수문자를 언더스코어로 대체
        import re
        name = re.sub(r'[^\w\s]', '_', name)  # 특수문자를 언더스코어로 변환
        name = re.sub(r'\s+', '_', name)      # 공백을 언더스코어로 변환
        
        # SQLite 예약어 처리
        reserved_words = ['add', 'all', 'alter', 'and', 'as', 'autoincrement', 'between', 'case', 'check', 'collate', 
                         'commit', 'constraint', 'create', 'default', 'deferrable', 'delete', 'distinct', 'drop', 
                         'else', 'escape', 'except', 'exists', 'foreign', 'from', 'group', 'having', 'if', 'in', 
                         'index', 'insert', 'intersect', 'into', 'is', 'isnull', 'join', 'limit', 'not', 'notnull', 
                         'null', 'on', 'or', 'order', 'primary', 'references', 'select', 'set', 'table', 'then', 
                         'to', 'transaction', 'union', 'unique', 'update', 'using', 'values', 'when', 'where']
        
        if name.lower() in reserved_words:
            name = name + '_col'
        
        # 숫자로 시작하는 경우 앞에 'col_' 추가
        if name and name[0].isdigit():
            name = 'col_' + name
            
        return name
    
    # SQLite 데이터베이스 연결
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    
    # 각 엑셀 설정에 대해 처리
    for table_name, config in excel_configs.items():
        log(f"테이블 '{table_name}' 처리 중...")
        
        # 엑셀 파일 열기
        try:
            log(f"엑셀 파일 '{config['path']}' 열기...")
            workbook = openpyxl.load_workbook(config['path'], read_only=True, data_only=True)
            sheet = workbook[config['sheet_name']]
            log(f"시트 '{config['sheet_name']}' 로드 완료")
        except Exception as e:
            log(f"엑셀 파일 '{config['path']}' 열기 실패: {e}")
            continue
        
        # 헤더 행 결정 (book2는 두 번째 행을 헤더로 사용)
        header_row_num = 1  # 기본값은 첫 번째 행
        if 'header_row' in config:
            header_row_num = config['header_row']
        
        # 헤더 행에서 컬럼 이름 가져오기
        header_row = next(sheet.iter_rows(min_row=header_row_num, max_row=header_row_num))
        column_names = []
        column_indices = []  # 유효한 컬럼의 인덱스를 저장
        column_name_counts = {}  # 중복 컬럼 이름 처리를 위한 딕셔너리
        
        for idx, cell in enumerate(header_row):
            col_name = clean_column_name(cell.value)
            
            # 빈 컬럼 이름은 무시
            if not col_name:
                continue
                
            # 중복 컬럼 이름 처리
            if col_name in column_name_counts:
                column_name_counts[col_name] += 1
                col_name = f"{col_name}_{column_name_counts[col_name]}"
            else:
                column_name_counts[col_name] = 0
                
            column_names.append(col_name)
            column_indices.append(idx)  # 유효한 컬럼의 인덱스 저장
        
        log(f"컬럼: {', '.join(column_names)}")
        
        # 테이블이 존재하면 삭제
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
        
        # 테이블 생성
        create_table_sql = f"CREATE TABLE {table_name} ({', '.join([f'{col} TEXT' for col in column_names])})"
        cursor.execute(create_table_sql)
        log(f"테이블 '{table_name}' 생성 완료")
        
        # 데이터 시작 행 결정
        data_start_row = header_row_num + 1
        
        # 데이터 삽입
        rows = []
        row_count = 0
        for row in sheet.iter_rows(min_row=data_start_row):  # 헤더 다음 행부터 시작
            row_data = []
            for idx in column_indices:  # 유효한 컬럼만 처리
                cell_value = row[idx].value
                row_data.append(str(cell_value) if cell_value is not None else "")
            
            if any(row_data):  # 빈 행은 건너뜀
                rows.append(row_data)
                row_count += 1
                
                # 1000행마다 데이터베이스에 삽입하고 로그 출력
                if row_count % 1000 == 0:
                    log(f"{row_count}행 처리 중...")
        
        # 데이터 일괄 삽입
        placeholders = ', '.join(['?' for _ in column_names])
        insert_sql = f"INSERT INTO {table_name} ({', '.join(column_names)}) VALUES ({placeholders})"
        cursor.executemany(insert_sql, rows)
        log(f"총 {len(rows)}행 삽입 완료")
        
        # 워크북 닫기
        workbook.close()
    
    # 변경사항 저장 및 연결 종료
    conn.commit()
    conn.close()
    
    log(f"데이터베이스 '{db_file}'에 성공적으로 데이터를 저장했습니다.")
    return True


class ExcelToSqliteApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel to SQLite 변환기")
        self.root.geometry("800x600")
        
        # 기본 설정
        self.excel_configs = {}
        self.db_file = "mgui2.db"
        
        # UI 구성
        self.create_widgets()
        
        # 기본 설정 로드
        self.load_default_config()
        
    def create_widgets(self):
        # 프레임 생성
        top_frame = ttk.Frame(self.root, padding="10")
        top_frame.pack(fill=tk.X)
        
        config_frame = ttk.LabelFrame(self.root, text="설정", padding="10")
        config_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        log_frame = ttk.LabelFrame(self.root, text="로그", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # 상단 버튼 및 검색 UI
        self.convert_btn = ttk.Button(top_frame, text="엑셀 -> SQLite 변환", command=self.convert)
        self.convert_btn.pack(side=tk.LEFT, padx=5)
        
        # 매핑SEQ 검색 UI
        ttk.Label(top_frame, text="매핑SEQ:").pack(side=tk.LEFT, padx=(20, 5))
        
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(top_frame, textvariable=self.search_var, width=15)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_entry.bind("<Return>", lambda event: self.search_mapping_seq())
        
        self.search_btn = ttk.Button(top_frame, text="Search", command=self.search_mapping_seq)
        self.search_btn.pack(side=tk.LEFT, padx=5)
        
        # 설정 표시 영역 - 높이를 2배로 늘림
        self.config_text = ScrolledText(config_frame, height=20)
        self.config_text.pack(fill=tk.BOTH, expand=True)
        
        # 로그 표시 영역 - 높이를 1/3로 줄임
        self.log_text = ScrolledText(log_frame, height=5)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
    def load_default_config(self):
        # 기본 설정 로드
        self.excel_configs = { 
            'book1': {
                'path': 'C:/work/doc/1.xlsx',
                'sheet_name': '1',
                'header_row': 1  # 첫 번째 행이 헤더
            },
            'book2': {
                'path': 'C:/work/doc/2.xlsx',
                'sheet_name': '1',
                'header_row': 2  # 두 번째 행이 헤더
            }
        }
        self.update_config_display()
        
    def update_config_display(self):
        # 설정 정보 표시
        self.config_text.delete(1.0, tk.END)
        self.config_text.insert(tk.END, f"데이터베이스 파일: {self.db_file}\n")
        
        for table_name, config in self.excel_configs.items():
            self.config_text.insert(tk.END, f"테이블: {table_name}\n")
            self.config_text.insert(tk.END, f"  엑셀 파일: {config['path']}\n")
            self.config_text.insert(tk.END, f"  시트 이름: {config['sheet_name']}\n")
            self.config_text.insert(tk.END, f"  헤더 행: {config.get('header_row', 1)}\n")
    
    def log(self, message):
        # 로그 메시지 추가
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)  # 스크롤을 가장 아래로 이동
        self.root.update()  # UI 업데이트
    
    def search_mapping_seq(self):
        # 매핑SEQ로 검색하는 함수
        mapping_seq = self.search_var.get().strip()
        if not mapping_seq:
            messagebox.showwarning("검색 오류", "매핑SEQ를 입력하세요.")
            return
        
        try:
            # SQLite 데이터베이스 연결
            if not os.path.exists(self.db_file):
                messagebox.showwarning("검색 오류", f"데이터베이스 파일({self.db_file})이 존재하지 않습니다. 먼저 변환을 실행하세요.")
                return
                
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            # book2 테이블 존재 여부 확인
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='book2'")
            if not cursor.fetchone():
                messagebox.showwarning("검색 오류", "book2 테이블이 존재하지 않습니다. 먼저 변환을 실행하세요.")
                conn.close()
                return
            
            # book1 테이블 존재 여부 확인
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='book1'")
            book1_exists = cursor.fetchone() is not None
            
            # book2 컬럼 정보 가져오기
            cursor.execute(f"PRAGMA table_info(book2)")
            book2_columns = [column[1] for column in cursor.fetchall()]
            
            # book1 컬럼 정보 가져오기
            book1_columns = []
            if book1_exists:
                cursor.execute(f"PRAGMA table_info(book1)")
                book1_columns = [column[1] for column in cursor.fetchall()]
            
            # 매핑SEQ 컬럼 존재 여부 확인
            if "매핑SEQ" not in book2_columns:
                # 컬럼명이 정리되었을 수 있으므로 가능한 변형 확인
                mapping_seq_col = None
                for col in book2_columns:
                    if col.lower() == "매핑seq" or col.lower() == "매핑_seq" or col.lower() == "매핑seq_" or col.lower() == "매핑_seq_":
                        mapping_seq_col = col
                        break
                
                if not mapping_seq_col:
                    messagebox.showwarning("검색 오류", "매핑SEQ 컬럼이 존재하지 않습니다.")
                    conn.close()
                    return
            else:
                mapping_seq_col = "매핑SEQ"
            
            # book1의 매핑SEQ 컬럼 찾기
            book1_mapping_seq_col = None
            if book1_exists:
                if "매핑SEQ" in book1_columns:
                    book1_mapping_seq_col = "매핑SEQ"
                else:
                    for col in book1_columns:
                        if col.lower() == "매핑seq" or col.lower() == "매핑_seq" or col.lower() == "매핑seq_" or col.lower() == "매핑_seq_":
                            book1_mapping_seq_col = col
                            break
            
            # book1의 인터페이스_명 컬럼 찾기
            interface_name_col = None
            if book1_exists and book1_mapping_seq_col:
                for col in book1_columns:
                    if "인터페이스" in col and "명" in col:
                        interface_name_col = col
                        break
            
            # book1의 I_F_Type 컬럼 찾기
            if_type_col = None
            if book1_exists and book1_mapping_seq_col:
                for col in book1_columns:
                    if col.lower() == "i_f_type" or col.lower() == "if_type" or col.lower() == "i_f_타입" or col.lower() == "if_타입":
                        if_type_col = col
                        break
            
            # book1의 Route정의 컬럼 찾기
            route_col = None
            if book1_exists and book1_mapping_seq_col:
                for col in book1_columns:
                    if "route" in col.lower() and "정의" in col:
                        route_col = col
                        break
            
            # GroupID와 EventID 컬럼 확인
            group_id_col = None
            event_id_col = None
            for col in book2_columns:
                if "group" in col.lower() and "id" in col.lower():
                    group_id_col = col
                if "event" in col.lower() and "id" in col.lower():
                    event_id_col = col
            
            # 송신 및 수신 관련 컬럼 찾기
            send_task_col = None
            send_qmgr_col = None
            recv_task_col = None
            recv_qmgr_col = None
            
            # SQL 관련 컬럼 찾기
            send_userid_col = None
            send_passwd_col = None
            send_db_col = None
            send_schema_adapter_col = None
            send_table_adapter_col = None
            
            recv_userid_col = None
            recv_passwd_col = None
            recv_db_col = None
            recv_schema_adapter_col = None
            recv_table_adapter_col = None
            
            for col in book2_columns:
                # 송신 관련 컬럼
                if "송신" in col and "업무" in col:
                    send_task_col = col
                if "송신" in col and "qmgr" in col.lower():
                    send_qmgr_col = col
                if "송신" in col and "userid" in col.lower():
                    send_userid_col = col
                if "송신" in col and ("passwd" in col.lower() or "password" in col.lower()):
                    send_passwd_col = col
                if "송신" in col and "db" in col.lower():
                    send_db_col = col
                if "송신" in col and "schema" in col.lower() and "adapter" in col.lower():
                    send_schema_adapter_col = col
                if "송신" in col and "table" in col.lower() and "adapter" in col.lower():
                    send_table_adapter_col = col
                
                # 수신 관련 컬럼
                if "수신" in col and "업무" in col:
                    recv_task_col = col
                if "수신" in col and "qmgr" in col.lower():
                    recv_qmgr_col = col
                if "수신" in col and "userid" in col.lower():
                    recv_userid_col = col
                if "수신" in col and ("passwd" in col.lower() or "password" in col.lower()):
                    recv_passwd_col = col
                if "수신" in col and "db" in col.lower():
                    recv_db_col = col
                if "수신" in col and "schema" in col.lower() and "adapter" in col.lower():
                    recv_schema_adapter_col = col
                if "수신" in col and "table" in col.lower() and "adapter" in col.lower():
                    recv_table_adapter_col = col
            
            # book2에서 검색 실행
            query = f"SELECT * FROM book2 WHERE {mapping_seq_col} = ?"
            cursor.execute(query, (mapping_seq,))
            book2_row = cursor.fetchone()
            
            if not book2_row:
                messagebox.showinfo("검색 결과", f"매핑SEQ '{mapping_seq}'에 해당하는 데이터가 없습니다.")
                conn.close()
                return
            
            # book1에서 검색 실행
            book1_row = None
            if book1_exists and book1_mapping_seq_col:
                query = f"SELECT * FROM book1 WHERE {book1_mapping_seq_col} = ?"
                cursor.execute(query, (mapping_seq,))
                book1_row = cursor.fetchone()
            
            # 결과 표시
            self.config_text.delete(1.0, tk.END)
            
            # 매핑SEQ 표시
            self.config_text.insert(tk.END, f"[매핑SEQ] {book2_row[book2_columns.index(mapping_seq_col)]}\n")
            
            # INTERFACE ID 표시 (GroupID.EventID)
            interface_id = ""
            if group_id_col and event_id_col:
                group_id = book2_row[book2_columns.index(group_id_col)]
                event_id = book2_row[book2_columns.index(event_id_col)]
                interface_id = f"{group_id}.{event_id}"
            
            self.config_text.insert(tk.END, f"[INTERFACE ID] {interface_id}\n")
            
            # 인터페이스 명 표시 (book1 테이블에서)
            if book1_row and interface_name_col:
                interface_name = book1_row[book1_columns.index(interface_name_col)]
                if interface_name:
                    self.config_text.insert(tk.END, f"[인터페이스 명] {interface_name}\n")
            
            # I_F_Type 표시 (book1 테이블에서)
            if book1_row and if_type_col:
                if_type = book1_row[book1_columns.index(if_type_col)]
                if if_type and str(if_type).strip():
                    self.config_text.insert(tk.END, f"[I_F_Type] {if_type}\n")
            
            # 송신 정보 표시
            send_info = ""
            if send_task_col:
                send_task = book2_row[book2_columns.index(send_task_col)]
                send_info = f"[송신_업무명] {send_task}"
            
            if send_qmgr_col:
                send_qmgr = book2_row[book2_columns.index(send_qmgr_col)]
                if send_info:
                    send_info += f"    [송신_QMGR명] {send_qmgr}"
                else:
                    send_info = f"[송신_QMGR명] {send_qmgr}"
            
            if send_info:
                self.config_text.insert(tk.END, f"\n{send_info}")
            
            # 수신 정보 표시
            recv_info = ""
            if recv_task_col:
                recv_task = book2_row[book2_columns.index(recv_task_col)]
                recv_info = f"[수신_업무명] {recv_task}"
            
            if recv_qmgr_col:
                recv_qmgr = book2_row[book2_columns.index(recv_qmgr_col)]
                if recv_info:
                    recv_info += f"    [수신_QMGR명] {recv_qmgr}"
                else:
                    recv_info = f"[수신_QMGR명] {recv_qmgr}"
            
            if recv_info:
                self.config_text.insert(tk.END, f"\n{recv_info}")
            
            # SQL 정보 추가
            # 송신 SQL
            send_sql = "\n[송신SQL]"
            
            send_userid = book2_row[book2_columns.index(send_userid_col)] if send_userid_col and book2_columns.index(send_userid_col) < len(book2_row) else "?"
            send_passwd = book2_row[book2_columns.index(send_passwd_col)] if send_passwd_col and book2_columns.index(send_passwd_col) < len(book2_row) else "?"
            send_db = book2_row[book2_columns.index(send_db_col)] if send_db_col and book2_columns.index(send_db_col) < len(book2_row) else "?"
            send_schema = book2_row[book2_columns.index(send_schema_adapter_col)] if send_schema_adapter_col and book2_columns.index(send_schema_adapter_col) < len(book2_row) else "?"
            send_table = book2_row[book2_columns.index(send_table_adapter_col)] if send_table_adapter_col and book2_columns.index(send_table_adapter_col) < len(book2_row) else "?"
            
            send_sql += f"\nsqlplus {send_userid}/{send_passwd}@{send_db}"
            send_sql += f"\nselect count(*) from {send_schema}.{send_table} where EAI_TRANSFER_DATE > sysdate-(1/12) and EAI_TRANSFER_FLAG='Y';"
            
            self.config_text.insert(tk.END, send_sql)
            
            # 수신 SQL
            recv_sql = "\n[수신SQL]"
            
            recv_userid = book2_row[book2_columns.index(recv_userid_col)] if recv_userid_col and book2_columns.index(recv_userid_col) < len(book2_row) else "?"
            recv_passwd = book2_row[book2_columns.index(recv_passwd_col)] if recv_passwd_col and book2_columns.index(recv_passwd_col) < len(book2_row) else "?"
            recv_db = book2_row[book2_columns.index(recv_db_col)] if recv_db_col and book2_columns.index(recv_db_col) < len(book2_row) else "?"
            recv_schema = book2_row[book2_columns.index(recv_schema_adapter_col)] if recv_schema_adapter_col and book2_columns.index(recv_schema_adapter_col) < len(book2_row) else "?"
            recv_table = book2_row[book2_columns.index(recv_table_adapter_col)] if recv_table_adapter_col and book2_columns.index(recv_table_adapter_col) < len(book2_row) else "?"
            
            recv_sql += f"\nsqlplus {recv_userid}/{recv_passwd}@{recv_db}"
            recv_sql += f"\nselect count(*) from {recv_schema}.{recv_table} where EAI_TRANSFER_DATE > sysdate-(1/12);"
            
            self.config_text.insert(tk.END, recv_sql)
            
            # Route정의 표시 (book1 테이블에서)
            if book1_row and route_col:
                route_def = book1_row[book1_columns.index(route_col)]
                if route_def:
                    self.config_text.insert(tk.END, f"\n[Route정의] {route_def}")
            
            conn.close()
            self.log(f"매핑SEQ '{mapping_seq}' 검색 완료")
            
        except Exception as e:
            self.log(f"검색 중 오류 발생: {e}")
            messagebox.showerror("검색 오류", f"검색 중 오류가 발생했습니다: {e}")
    
    def convert(self):
        # 변환 시작
        self.log("변환 시작...")
        self.convert_btn.config(state=tk.DISABLED)
        
        try:
            # 변환 실행
            success = excel_to_sqlite(self.excel_configs, self.db_file, self.log)
            
            if success:
                messagebox.showinfo("완료", f"변환이 완료되었습니다.\n데이터베이스 파일: {self.db_file}")
            
        except Exception as e:
            self.log(f"오류 발생: {e}")
            messagebox.showerror("오류", f"변환 중 오류가 발생했습니다: {e}")
        
        finally:
            self.convert_btn.config(state=tk.NORMAL)


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelToSqliteApp(root)
    root.mainloop()